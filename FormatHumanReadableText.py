import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


def process_csv(input_csv_path):
    """
    Reads the input CSV file and prepares a DataFrame with the required columns.
    Groups responses by 'hateSpeech', sorts them by 'rrScore', and collects them.
    The 'hateScore' and 'userEnteredResponse' columns are left blank.
    """
    # Read the CSV file
    df = pd.read_csv(input_csv_path)

    # Group by 'hateSpeech'
    grouped = df.groupby('hateSpeech')

    # Prepare the list to collect rows
    data = []

    # For each group, collect responses sorted by 'rrScore'
    for hate_speech, group in grouped:
        # Sort responses by 'rrScore' in descending order
        sorted_group = group.sort_values(by='rrScore', ascending=False)

        # Get the top 4 responses
        responses = sorted_group['response'].tolist()[:4]

        # Pad the responses list to ensure it has 4 elements
        responses += [''] * (4 - len(responses))

        # Create a new row
        row = {
            'hatespeech': hate_speech,
            'hateScore': '',  # Leave blank
            'userEnteredResponse': '',  # Leave blank
            'generatedResponse1': responses[0],
            'generatedResponse2': responses[1],
            'generatedResponse3': responses[2],
            'generatedResponse4': responses[3]
        }
        data.append(row)

    # Create a new DataFrame
    new_df = pd.DataFrame(data)

    return new_df


def adjust_excel_formatting(excel_path):
    """
    Adjusts the Excel file formatting so that cells are taller and text wraps.
    """
    wb = load_workbook(excel_path)
    ws = wb.active

    # Set alignment and wrap text for all cells
    alignment = Alignment(wrap_text=True, vertical='top')

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = alignment

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column index
        column_letter = get_column_letter(column)
        for cell in col:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width if adjusted_width < 50 else 50  # Limit max width

    # Increase row heights significantly
    for row in ws.iter_rows():
        max_lines = 0
        for cell in row:
            if cell.value:
                lines = str(cell.value).count('\n') + 1
                if lines > max_lines:
                    max_lines = lines
        row_number = row[0].row
        ws.row_dimensions[row_number].height = max_lines * 20  # Increase multiplier as needed

    wb.save(excel_path)


def csv_to_formatted_excel(input_csv_path, output_excel_path):
    """
    Main function to convert CSV to formatted Excel.
    """
    # Process CSV and get DataFrame
    df = process_csv(input_csv_path)

    # Write DataFrame to Excel
    df.to_excel(output_excel_path, index=False)

    # Adjust Excel formatting
    adjust_excel_formatting(output_excel_path)


if __name__ == '__main__':
    csv_to_formatted_excel('advercialModel/rankedOutputToBeGraded/4-5.csv', 'humanReadableCS/4-5.xlsx')
    csv_to_formatted_excel('advercialModel/rankedOutputToBeGraded/7-8.csv', 'humanReadableCS/7-8.xlsx')
